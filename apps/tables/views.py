import json
import csv
import os
import requests
import re
from django.http import HttpResponse
from django.shortcuts import render, redirect, get_object_or_404
#from apps.tables.forms import DeviceForm
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from apps.tables.utils import (
    device_filter, global_filter, software_filter,
    software_assets_filter, same_key_filter, 
    unique_filter_func, common_date_filter,
    common_float_filter, common_integer_filter,
    common_count_filter, common_unique_filter,
    create_hide_show_filter, get_model_fields
)
from django.conf import settings
from apps.tables.models import UserFilter, PageItems, HideShowFilter, IntRangeFilter, FloatRangeFilter, ModelChoices, DateRangeFilter, TaskStatus
from django.http import HttpResponse, FileResponse, Http404
from django.views import View
from django.db import models
from django.urls import reverse, resolve
from django.http import JsonResponse
from datetime import datetime
from django.utils import timezone
from loader.models import InstantUpload, TypeChoices, DynamicQuery, SAPApi
from django.contrib.contenttypes.models import ContentType
from dateutil.parser import isoparse
from django.db.models import OuterRef, Subquery, F, Window, Count
from django.db.models.functions import RowNumber
from django.db import connection, connections
from apps.common.models import SavedFilter
from django.contrib.auth import get_user_model
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from home.models import ColumnOrder, JoinModel, PyFunction, PyFunctionPrompt
from django.views.decorators.csrf import csrf_exempt
try:
    from pgvector.django import VectorField
    PGVECTOR_AVAILABLE = True
except ImportError:
    PGVECTOR_AVAILABLE = False

User = get_user_model()

# Create your views here.

def get_user_id(request):
    if request.user.is_authenticated:
        return request.user.pk
    else:
        return -1

def create_device_filter(request):
    if request.method == "POST":
        from_dates = request.POST.getlist('from_date')
        to_dates = request.POST.getlist('to_date')
        keys = request.POST.getlist('key_date_range')

        if len(from_dates) == len(to_dates) == len(keys):
            for key, from_date, to_date in zip(keys, from_dates, to_dates):
                existing_record = DateRangeFilter.objects.filter(
                    userID=get_user_id(request),
                    parent=ModelChoices.DEVICES,
                    key=key,
                    from_date=from_date,
                    to_date=to_date
                ).first()

                if existing_record:
                    existing_record.from_date = from_date
                    existing_record.to_date = to_date
                    existing_record.save()
                else:
                    DateRangeFilter.objects.create(
                        userID=get_user_id(request),
                        parent=ModelChoices.DEVICES,
                        key=key,
                        from_date=from_date,
                        to_date=to_date
                    )

        values = request.POST.getlist('value')
        value_change = request.POST.getlist('value_change')

        keys = request.POST.getlist('key')
        keys_change = request.POST.getlist('key_change')
        change_field_id = request.POST.getlist('change_field_id')

        for i in range(len(values)):
            key = keys[i]
            value = values[i]

            UserFilter.objects.create(
                userID=get_user_id(request),
                parent=ModelChoices.DEVICES,
                key=key,
                value=value
            )
        
        for i in range(len(value_change)):
            key = keys_change[i]
            value = value_change[i]
            id = change_field_id[i]

            existing_record = UserFilter.objects.get(
                userID=get_user_id(request),
                parent=ModelChoices.DEVICES,
                id=id
            )

            if existing_record:
                existing_record.value = value
                existing_record.key = key
                existing_record.save()

        return redirect(reverse('list_of_devices'))



def create_device_page_items(request):
    if request.method == 'POST':
        items = request.POST.get('items')
        page_items, created = PageItems.objects.update_or_create(
            userID=get_user_id(request),
            parent=ModelChoices.DEVICES,
            defaults={'items_per_page':items}
        )
        return redirect(reverse('list_of_devices'))

def create_device_hide_show_filter(request):
    if request.method == "POST":
        data_str = list(request.POST.keys())[0]
        data = json.loads(data_str)


        HideShowFilter.objects.update_or_create(
            userID=get_user_id(request),
            parent=ModelChoices.DEVICES,
            key=data.get('key'),
            defaults={'value': data.get('value')}
        )

        response_data = {'message': 'Model updated successfully'}
        return JsonResponse(response_data)

    return JsonResponse({'error': 'Invalid request'}, status=400)
    

def delete_device_filter(request, id):
    filter_instance = UserFilter.objects.get(id=id, userID=get_user_id(request), parent=ModelChoices.DEVICES)
    filter_instance.delete()
    return redirect(reverse('list_of_devices'))

def delete_daterange_filter(request, id):
    date_filter = DateRangeFilter.objects.get(id=id, userID=get_user_id(request), parent=ModelChoices.DEVICES)
    date_filter.delete()
    return redirect(reverse('list_of_devices'))

def list_of_devices(request):
    db_field_names = [field.name for field in Devices._meta.get_fields() if not field.is_relation if not '_Q_' in field.name]
    
    # Snapshot
    latest_snapshot = ''
    summary = None
    snapshot_filter = {}
    try:
        content_type = ContentType.objects.get(model='devices')
        snapshots = InstantUpload.objects.filter(content_type=content_type).order_by('-created_at')
        snapshot = request.GET.get('snapshot')
        latest_snapshot = snapshots.latest('created_at')
        if latest_snapshot and not snapshot == 'all':
            snapshot_filter['loader_instance'] = latest_snapshot.pk
        
        if snapshot and not snapshot == 'all':
            summary = InstantUpload.objects.get(id=snapshot)
            snapshot_filter['loader_instance'] = summary.pk

    except:
        pass


    field_names = []

    for field_name in db_field_names:
        fields, created = HideShowFilter.objects.get_or_create(key=field_name, userID=get_user_id(request), parent=ModelChoices.DEVICES)
        field_names.append(fields)

    page_items = PageItems.objects.filter(userID=get_user_id(request), parent=ModelChoices.DEVICES).last()
    items = 25
    if page_items:
        items = page_items.items_per_page

    filter_string = {}
    # Pre-FILTERS
    #filter_string['MatchedTitle__icontains'] = 'hdx'

    filter_instance = UserFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.DEVICES)

    combined_q_objects, unique_filter = same_key_filter(filter_instance)

    date_filter_instance = DateRangeFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.DEVICES)
    for date_filter in date_filter_instance:
        from_date = datetime.strptime(date_filter.from_date.strftime('%Y-%m-%d'), "%Y-%m-%d").timestamp()
        to_date = datetime.strptime((date_filter.to_date + timezone.timedelta(days=1)).strftime('%Y-%m-%d'), "%Y-%m-%d").timestamp()
        filter_string[f'{date_filter.key}__range'] = [from_date, to_date]


    order_by = request.GET.get('order_by', 'ID')
    queryset = Devices.objects.filter(**filter_string).filter(combined_q_objects).filter(**snapshot_filter).order_by(order_by)

    if unique_filter:
        subquery = queryset.filter(**{unique_filter: OuterRef(unique_filter)}).order_by('pk').values('pk')[:1]
        queryset = queryset.filter(pk__in=Subquery(subquery))

    device_list = device_filter(request, queryset, db_field_names)
    # form = DeviceForm()

    page = request.GET.get('page', 1)
    paginator = Paginator(device_list, items)
    
    try:
        devices = paginator.page(page)
    except PageNotAnInteger:
        return redirect('list_of_devices')
    except EmptyPage:
        return redirect('list_of_devices') 

    # if request.method == 'POST':
    #     form = DeviceForm(request.POST)
    #     if form.is_valid():
    #         return post_request_handling(request, form)

    read_only_fields = ('ID', 'PrimaryUserName', 'DeviceLastConnUtc', )
    pre_column = ('ID', 'ImportDate', )
    date_picker_fields = ('ImportDate', 'DeviceLastConnUtc', )

    COMBOS = {}
    COMBOS['DeviceType'] = ['Virtual Machine', 'PC', 'Server' ]
    COMBOS['OperatingSystem'] = ['Windows', 'Linux', 'MAC']

    context = {
        'segment'  : 'list_of_devices',
        'parent'   : 'user_compliance',
        # 'form'     : form,
        'devices' : devices,
        'field_names': field_names,
        'db_field_names': db_field_names,
        'filter_instance': filter_instance,
        'date_filter_instance': DateRangeFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.DEVICES),
        'items': items,
        'read_only_fields': read_only_fields,
        'total_items': Devices.objects.count(),
        'pre_column': pre_column,
        'COMBOS': COMBOS,
        'date_picker_fields': date_picker_fields,
        'current_time': datetime.now().strftime("%Y-%m-%d"),
        'page_items': page_items,

        # snapshots
        'snapshots': snapshots,
        'latest_snapshot': latest_snapshot,
        'selected_snapshot': summary
    }
    
    return render(request, 'apps/list_of_devices.html', context)



@login_required(login_url='/users/signin/')
def create_device(request):
    date_picker_fields = ('ImportDate', 'DeviceLastConnUtc', )

    if request.method == 'POST':
        device_data = {}
        for attribute, value in request.POST.items():
            if attribute == 'csrfmiddlewaretoken':
                continue


            if attribute not in date_picker_fields:
                if attribute in ['IsLicenseRequired']:
                    device_data[attribute] = value if value else 0
                else:
                    device_data[attribute] = value if value else ''
            else:
                unix_time = ''
                if value:
                    unix_time = datetime.strptime(value, "%Y-%m-%dT%H:%M").timestamp()
                device_data[attribute] = unix_time

        Devices.objects.create(**device_data)

    return redirect(request.META.get('HTTP_REFERER'))

@login_required(login_url='/users/signin/')
def delete_device(request, id):
    device = Devices.objects.get(ID=id)
    device.delete()
    return redirect(request.META.get('HTTP_REFERER'))


@login_required(login_url='/users/signin/')
def update_device(request, id):
    product = Devices.objects.get(ID=id)
    date_picker_fields = ('ImportDate', 'DeviceLastConnUtc', )

    if request.method == 'POST':
        for attribute, value in request.POST.items():
            if attribute == 'csrfmiddlewaretoken':
                continue
            

            if not attribute in date_picker_fields:
                setattr(product, attribute, value)
            else:
                unix_time = setattr(product, attribute, value)
                if value:
                    unix_time = datetime.strptime(value, "%Y-%m-%dT%H:%M").timestamp()
                setattr(product, attribute, unix_time)

        product.save()

    return redirect(request.META.get('HTTP_REFERER'))


# Export as CSV
class ExportCSVView(View):
    def get(self, request):
        db_field_names = [field.name for field in Devices._meta.get_fields() if not field.is_relation]
        
        # Snapshot
        latest_snapshot = ''
        summary = None
        snapshot_filter = {}
        try:
            content_type = ContentType.objects.get(model='devices')
            snapshots = InstantUpload.objects.filter(content_type=content_type).order_by('-created_at')
            snapshot = request.GET.get('snapshot')
            latest_snapshot = snapshots.latest('created_at')
            if latest_snapshot and not snapshot == 'all':
                snapshot_filter['loader_instance'] = latest_snapshot.pk
            
            if snapshot and not snapshot == 'all':
                summary = InstantUpload.objects.get(id=snapshot)
                snapshot_filter['loader_instance'] = summary.pk

        except:
            pass
        
        fields = []
        pre_column = ('ID', 'ImportDate', )
        show_fields = HideShowFilter.objects.filter(value=False, userID=get_user_id(request), parent=ModelChoices.DEVICES)
        for field in show_fields:
            if field.key not in pre_column:
                fields.append(field.key)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="devices.csv"'

        writer = csv.writer(response)
        writer.writerow(fields) 


        filter_string = {}
        filter_instance = UserFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.DEVICES)
        # for filter_data in filter_instance:
        #     filter_string[f'{filter_data.key}__icontains'] = filter_data.value
        combined_q_objects, unique_filter = same_key_filter(filter_instance)

        order_by = request.GET.get('order_by', 'ID')
        queryset = Devices.objects.filter(**filter_string).filter(combined_q_objects).filter(**snapshot_filter).order_by(order_by)

        if unique_filter:
            subquery = queryset.filter(**{unique_filter: OuterRef(unique_filter)}).order_by('pk').values('pk')[:1]
            queryset = queryset.filter(pk__in=Subquery(subquery))

        devices = device_filter(request, queryset, db_field_names)

        for device in devices:
            row_data = [getattr(device, field) for field in fields]
            writer.writerow(row_data)

        return response


from reportlab.pdfgen import canvas
from io import BytesIO
# Export as PDF
class ExportPDFView(View):
    def get(self, request):
        buffer = BytesIO()
        response_pdf = HttpResponse(content_type='application/pdf')
        response_pdf['Content-Disposition'] = 'attachment; filename="devices.pdf"'

        pdf = canvas.Canvas(buffer)

        page_number = request.GET.get('page', 1)
        items_per_page = 25

        devices = Devices.objects.all()
        paginator = Paginator(devices, items_per_page)
        current_page_devices = paginator.get_page(page_number)

        y_position = 800

        for device in current_page_devices:
            pdf.drawString(100, y_position, f"ID: {device.ID}, DeviceName: {device.DeviceName}, OperatingSystem: {device.OperatingSystem}")
            y_position -= 20

        pdf.save()

        buffer.seek(0)
        response_pdf.write(buffer.getvalue())

        return response_pdf
    






# Many User DT

def user_devices(request):
    # Top DT
    user_filter_string = software_assets_filter(['PrimaryUserName'], request.GET.get('user_search'))
    users = Devices.objects.filter(user_filter_string).values('PrimaryUserName').distinct()

    # Bottom DT
    db_field_names = [field.name for field in Devices._meta.get_fields() if not field.is_relation if not '_Q_' in field.name]

    page_items = PageItems.objects.filter(userID=get_user_id(request), parent=ModelChoices.DEVICES).last()
    items = 25
    if page_items:
        items = page_items.items_per_page
    

    device_filter_string = software_assets_filter(db_field_names, request.GET.get('device_search'))
    order_by = request.GET.get('order_by', 'ID')
    queryset = Devices.objects.filter(device_filter_string).order_by(order_by)

    exclude_columns = ('MatchedVendor', 'MatchedTitle', 'MatchedVersion', 'OperatingSystem')

    primary_usernames = UserDeviceAssets.objects.values_list('primary_username', flat=True)
    selected_rows = []
    if request.GET.get('user'):
        selected_rows = UserDeviceAssets.objects.filter(primary_username=request.GET.get('user')).values_list('dt_rows__ID', flat=True)


    device_list = device_filter(request, queryset, db_field_names)

    page = request.GET.get('page', 1)
    paginator = Paginator(device_list, items)
    
    try:
        devices = paginator.page(page)
    except PageNotAnInteger:
        return redirect('list_of_devices')
    except EmptyPage:
        return redirect('list_of_devices') 
    
    selected_devices = []
    remaining_devices = []

    for device in sorted(Devices.objects.all(), key=lambda x: x.ID):
        if device.ID in selected_rows:
            selected_devices.append(device)

    for device in devices:
        if device.ID not in selected_rows:
            remaining_devices.append(device)


    context = {
        'segment': 'list_of_ham_devices',
        'parent': 'software_assets',
        'users': users,
        'devices': devices,
        'selected_devices': selected_devices,
        'remaining_devices': remaining_devices,
        'db_field_names': db_field_names,
        'exclude_columns': exclude_columns,
        'db_top_field_names': ['PrimaryUserName'],
        'total_items': Devices.objects.count(),
        'primary_usernames': primary_usernames,
        'selected_rows': selected_rows
    }
    return render(request, 'apps/user_devices.html', context)



def save_selected_dt(request):
    if request.method == 'POST':
        username = request.POST.get('primary_user', request.GET.get('user'))
        devices = request.POST.getlist('devices')
        assets, created = UserDeviceAssets.objects.update_or_create(primary_username=username)
        assets.dt_rows.set(Devices.objects.filter(pk__in=devices))

    return redirect(reverse('user_devices'))




####################
# Favorite
####################
from apps.tables.models import Favorite, ActionStatus
from django.contrib.contenttypes.models import ContentType
from apps.common.models import SavedFilter, FieldType
import json


def create_favorite_filter(request, favorite_id):
    favorite = Favorite.objects.get(id=favorite_id)
    content_type = ContentType.objects.get(app_label=favorite.content_type.app_label, model=favorite.content_type.model)
    model = content_type.model_class()

    field_type_map = {
        'date': getattr(model, 'date_fields_to_convert', []),
        'int': getattr(model, 'integer_fields', []),
        'float': getattr(model, 'float_fields', [])
    }

    def get_field_type(value):
        for f_type, fields in field_type_map.items():
            if value in fields:
                return getattr(FieldType, f_type.upper())
        return FieldType.TEXT

    def get_boolean_field(name, index):
        return request.POST.get(f'{name}_{index}') == 'true'

    existing_count_filter = SavedFilter.objects.filter(
        userID=request.user.id, parent=ModelChoices.FAVORITE, is_count=True, favorite_id=favorite.id
    ).first()
    for key, value in request.POST.items():
        if key.startswith('field_name_') and value:
            index = key.split('_')[-1]
            field_name = value
            field_type = get_field_type(value)

            if field_type == FieldType.TEXT:
                value_start = request.POST.get(f'value_start_{index}') or ''
                value_end = request.POST.get(f'value_end_{index}') or ''
            else:
                value_start = request.POST.get(f'value_start_{index}') or None
                value_end = request.POST.get(f'value_end_{index}') or None

            is_null = get_boolean_field('is_null', index)
            is_not_null = get_boolean_field('is_not_null', index)
            is_unique = get_boolean_field('is_unique', index)
            is_count = get_boolean_field('is_count', index)

            filter_id = request.POST.get(f'filter_id_{index}')
            if filter_id in [None, '', 'null']:
                filter_id = None
            
            if is_count:
                if existing_count_filter and (not filter_id or str(existing_count_filter.id) != filter_id):
                    is_count = False

            filter_data = {
                'userID': request.user.id,
                'parent': ModelChoices.FAVORITE,
                'field_name': field_name,
                'field_type': field_type,
                'value_start': value_start,
                'value_end': value_end,
                'is_null': is_null,
                'is_not_null': is_not_null,
                'is_unique': is_unique,
                'is_count': is_count,
                'favorite_id': favorite.id
            }

            if filter_id:
                SavedFilter.objects.filter(pk=filter_id).update(**filter_data)
                saved_filter = SavedFilter.objects.get(pk=filter_id)
            else:
                saved_filter = SavedFilter.objects.create(**filter_data)
            
            if not favorite.saved_filters.filter(pk=saved_filter.pk).exists():
                favorite.saved_filters.add(saved_filter)
    
    return redirect(request.META.get('HTTP_REFERER'))


def create_favorite_page_items(request, favorite_id):
    favorite = Favorite.objects.get(id=favorite_id)
    if request.method == 'POST':
        items = request.POST.get('items')
        page_items, created = PageItems.objects.update_or_create(
            userID=get_user_id(request),
            parent=ModelChoices.FAVORITE,
            favorite_id=favorite.id,
            defaults={'items_per_page':items}
        )
        favorite.page_items = page_items
        favorite.save()

        return redirect(reverse('list_of_devices'))

def create_favorite_hide_show_filter(request, favorite_id):
    if request.method == "POST":
        data_list = json.loads(request.body)
        favorite = Favorite.objects.get(id=favorite_id)

        for data in data_list:
            hide_show_filter, created = HideShowFilter.objects.update_or_create(
                userID=get_user_id(request),
                parent=ModelChoices.FAVORITE,
                key=data.get('key'),
                favorite_id=favorite.id,
                defaults={'value': data.get('value')}
            )
            favorite.hide_show_filters.add(hide_show_filter)

        response_data = {'message': 'Filters updated successfully'}
        return JsonResponse(response_data)

    return JsonResponse({'error': 'Invalid request'}, status=400)

def delete_favorite_filter(request):
    if request.method == 'POST':
        filter_id = request.POST.get('filter_id')
        try:
            # Delete the filter from the database
            filter_to_delete = SavedFilter.objects.get(id=filter_id, userID=request.user.id)
            filter_to_delete.delete()
            return JsonResponse({'success': True})
        except SavedFilter.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Filter not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

def delete_favorite_date_range_filter(request, id):
    filter_instance = DateRangeFilter.objects.get(id=id, userID=get_user_id(request), parent=ModelChoices.FAVORITE)
    filter_instance.delete()

    return redirect(request.META.get('HTTP_REFERER'))

def delete_favorite_int_range_filter(request, id):
    filter_instance = IntRangeFilter.objects.get(id=id, userID=get_user_id(request), parent=ModelChoices.FAVORITE)
    filter_instance.delete()

    return redirect(request.META.get('HTTP_REFERER'))

def delete_favorite_float_range_filter(request, id):
    filter_instance = FloatRangeFilter.objects.get(id=id, userID=get_user_id(request), parent=ModelChoices.FAVORITE)
    filter_instance.delete()

    return redirect(request.META.get('HTTP_REFERER'))



@login_required(login_url='/users/signin/')
def add_to_favorite(request, model_name, model_choice):
    if request.method == 'POST':
        content_type = ContentType.objects.get(model=model_name.lower())
        img_loader_id = request.POST.get("img_loader_id")
        user_filter_qs = UserFilter.objects.filter(
            userID=get_user_id(request), parent=getattr(ModelChoices, model_choice),
            **({"img_loader_id": img_loader_id} if img_loader_id else {})
        )
        date_range_filter_qs = DateRangeFilter.objects.filter(
            userID=get_user_id(request), parent=getattr(ModelChoices, model_choice),
            **({"img_loader_id": img_loader_id} if img_loader_id else {})
        )

        int_range_filter_qs = IntRangeFilter.objects.filter(
            userID=get_user_id(request), parent=getattr(ModelChoices, model_choice),
            **({"img_loader_id": img_loader_id} if img_loader_id else {})
        )

        float_range_filter_qs = FloatRangeFilter.objects.filter(
            userID=get_user_id(request), parent=getattr(ModelChoices, model_choice),
            **({"img_loader_id": img_loader_id} if img_loader_id else {})
        )

        saved_filter_qs = SavedFilter.objects.filter(
            userID=get_user_id(request), parent=getattr(ModelChoices, model_choice),
            **({"img_loader_id": img_loader_id} if img_loader_id else {})
        )

        hide_show_filter_qs = HideShowFilter.objects.filter(
            userID=get_user_id(request), parent=getattr(ModelChoices, model_choice),
            **({"img_loader_id": img_loader_id} if img_loader_id else {})
        )

        page_items_qs = PageItems.objects.filter(
            userID=get_user_id(request), parent=getattr(ModelChoices, model_choice),
            **({"img_loader_id": img_loader_id} if img_loader_id else {})
        ).last()
        
        fav_name = request.POST.get('fav_name')

        favorite = Favorite.objects.create(
            user=request.user,
            name=fav_name,
            content_type=content_type,
            model_choices=model_choice,
            search=request.POST.get('search_fields'),
            pre_columns=request.POST.get('pre_columns'),
            pre_filters=request.POST.get('pre_filters'),
            richtext_fields=request.POST.get('richtext_fields'),
            order_by=request.POST.get('order_by'),
            snapshot=request.POST.get('snapshot') if request.POST.get('snapshot') else 'latest',
            created_at=timezone.now(),
            updated_at=timezone.now()
        )

        if request.POST.get('query_snapshot'):
            favorite.query_snapshot = request.POST.get('query_snapshot')
        
        if request.POST.get('is_dynamic_query') == 'true':
            favorite.is_dynamic_query = True

        if page_items_qs:
            page_items = PageItems.objects.create(
                userID=page_items_qs.userID,
                parent=ModelChoices.FAVORITE,
                items_per_page=page_items_qs.items_per_page,
                favorite_id=favorite.pk
            )

            favorite.page_items = page_items

        for user_filter in user_filter_qs:
            user_filters = UserFilter.objects.create(
                userID=user_filter.userID,
                parent=ModelChoices.FAVORITE,
                key=user_filter.key,
                value=user_filter.value,
                favorite_id=favorite.pk
            )

            favorite.user_filters.add(user_filters)
    
        for date_range_filter in date_range_filter_qs:
            date_range_filters = DateRangeFilter.objects.create(
                userID=date_range_filter.userID,
                parent=ModelChoices.FAVORITE,
                from_date=date_range_filter.from_date,
                to_date=date_range_filter.to_date,
                key=date_range_filter.key,
                favorite_id=favorite.pk
            )

            favorite.date_range_filters.add(date_range_filters)
        
        for int_range_filter in int_range_filter_qs:
            int_range_filters = IntRangeFilter.objects.create(
                userID=int_range_filter.userID,
                parent=ModelChoices.FAVORITE,
                from_number=int_range_filter.from_number,
                to_number=int_range_filter.to_number,
                key=int_range_filter.key,
                favorite_id=favorite.pk
            )

            favorite.int_range_filters.add(int_range_filters)
        
        for float_range_filter in float_range_filter_qs:
            float_range_filters = FloatRangeFilter.objects.create(
                userID=float_range_filter.userID,
                parent=ModelChoices.FAVORITE,
                from_float_number=float_range_filter.from_float_number,
                to_float_number=float_range_filter.to_float_number,
                key=float_range_filter.key,
                favorite_id=favorite.pk
            )

            favorite.float_range_filters.add(float_range_filters)
        
        for saved_filter in saved_filter_qs:
            saved_filters = SavedFilter.objects.create(
                userID=saved_filter.userID,
                parent=ModelChoices.FAVORITE,
                field_name=saved_filter.field_name,
                field_type=saved_filter.field_type,
                value_start=saved_filter.value_start,
                value_end=saved_filter.value_end,
                is_null=saved_filter.is_null,
                is_not_null=saved_filter.is_not_null,
                is_unique=saved_filter.is_unique,
                is_count=saved_filter.is_count,
                favorite_id=favorite.pk
            )
            favorite.saved_filters.add(saved_filters)

        for hide_show_filter in hide_show_filter_qs:
            hide_show_filters = HideShowFilter.objects.create(
                userID=hide_show_filter.userID,
                parent=ModelChoices.FAVORITE,
                key=hide_show_filter.key,
                value=hide_show_filter.value,
                favorite_id=favorite.pk
            )

            favorite.hide_show_filters.add(hide_show_filters)
        
        favorite.save()

    return redirect(request.META.get('HTTP_REFERER'))


@login_required(login_url='/users/signin/')
def add_favorite_from_favorite(request, favorite_id):
    favorite = Favorite.objects.get(id=favorite_id)

    if request.method == 'POST':
        if request.POST.get('search_fields'):
            favorite.search = request.POST.get('search_fields')
        
        if request.POST.get('order_by') :
            favorite.order_by = request.POST.get('order_by') 

        if request.POST.get('snapshot'):
            favorite.snapshot = request.POST.get('snapshot') 

        if favorite.page_items:
            page_items = PageItems.objects.get(
                userID=favorite.page_items.userID,
                parent=ModelChoices.FAVORITE,
                pk=favorite.page_items.pk
            )

            favorite.page_items = page_items

        if favorite.user_filters.all():
            for user_filter in favorite.user_filters.all():
                user_filters = UserFilter.objects.get(
                    userID=user_filter.userID,
                    parent=ModelChoices.FAVORITE,
                    pk=user_filter.pk
                )

                favorite.user_filters.add(user_filters)

        if favorite.date_range_filters.all():
            for date_range_filter in favorite.date_range_filters.all():
                date_range_filters = DateRangeFilter.objects.get(
                    userID=date_range_filter.userID,
                    parent=ModelChoices.FAVORITE,
                    pk=date_range_filter.pk
                )

                favorite.date_range_filters.add(date_range_filters)

        if favorite.int_range_filters.all():
            for int_range_filter in favorite.int_range_filters.all():
                int_range_filters = IntRangeFilter.objects.get(
                    userID=int_range_filter.userID,
                    parent=ModelChoices.FAVORITE,
                    pk=int_range_filter.pk
                )

                favorite.int_range_filters.add(int_range_filters)
        

        if favorite.float_range_filters.all():
            for float_range_filter in favorite.float_range_filters.all():
                float_range_filters = FloatRangeFilter.objects.get(
                    userID=float_range_filter.userID,
                    parent=ModelChoices.FAVORITE,
                    pk=float_range_filter.pk
                )

                favorite.float_range_filters.add(float_range_filters)
        
        if favorite.hide_show_filters.all():
            for hide_show_filter in favorite.hide_show_filters.all():
                hide_show_filters = HideShowFilter.objects.get(
                    userID=hide_show_filter.userID,
                    parent=ModelChoices.FAVORITE,
                    pk=hide_show_filter.pk
                )

                favorite.hide_show_filters.add(hide_show_filters)
        
        favorite.updated_at = timezone.now()
        favorite.save()

    return redirect(request.META.get('HTTP_REFERER'))

def favorite_details(request, id):
    favorite = Favorite.objects.get(id=id)
    content_type = ContentType.objects.get(app_label=favorite.content_type.app_label, model=favorite.content_type.model)
    db_field_names = [field.name for field in content_type.model_class()._meta.get_fields() if not field.is_relation and not '_Q_' in field.name]
    favorite_model = content_type.model_class()

    # Snapshot
    latest_snapshot = ''
    summary = None
    snapshot_filter = {}

    if not favorite.is_dynamic_query:
        try:
            content_type = ContentType.objects.get(model=favorite_model.__name__.lower())
            snapshots = InstantUpload.objects.filter(content_type=content_type).order_by('-created_at')
            snapshot = request.GET.get('snapshot')

            if snapshot and not snapshot == 'all':
                summary = InstantUpload.objects.get(id=snapshot)
                snapshot_filter['loader_instance'] = summary.pk
            
            elif snapshot and snapshot == 'all':
                snapshot_filter= {}
            
            elif favorite.snapshot == 'latest':
                latest_snapshot = snapshots.latest('created_at')
                snapshot_filter['loader_instance'] = latest_snapshot.pk
            
            elif favorite.snapshot and not favorite.snapshot == 'all':
                summary = InstantUpload.objects.get(id=int(favorite.snapshot))
                snapshot_filter['loader_instance'] = int(favorite.snapshot)
            
            elif favorite.snapshot and favorite.snapshot == 'all':
                snapshot_filter= {}
            
        except:
            pass
    
    else:
        snapshots = favorite_model.objects.exclude(snapshot=None).values('snapshot').distinct()
    
    items = 25
    if favorite.page_items:
        items = favorite.page_items.items_per_page

    filter_string = {}
    combined_q_objects, user_unique_filter, user_query_conditions, user_count_filters = same_key_filter(favorite.saved_filters.filter(field_type=FieldType.TEXT), return_count_filters=True)

    pre_filters = {}
    if favorite.pre_filters:
        pre_filters = eval(favorite.pre_filters)
    
    if favorite.is_dynamic_query:
        query_snapshot = favorite.query_snapshot
        if request.GET.get('query_snapshot'):
            query_snapshot = request.GET.get('query_snapshot')
        else:
            query_snapshot = favorite.query_snapshot
        

        if query_snapshot and query_snapshot != 'all':
            parsed_datetime = isoparse(query_snapshot)
            snapshot_filter['snapshot'] = parsed_datetime
    
    # for date range
    date_string, date_unique_filter, date_query_conditions, date_count_filters = common_date_filter(favorite.saved_filters.filter(field_type=FieldType.DATE), return_count_filters=True)
    filter_string.update(date_string)
    
    # for integer range
    int_string, int_unique_filter, int_query_conditions, int_count_filters = common_integer_filter(favorite.saved_filters.filter(field_type=FieldType.INT), return_count_filters=True)
    filter_string.update(int_string)

    # for float range
    float_string, float_unique_filter, float_query_conditions, float_count_filters = common_float_filter(favorite.saved_filters.filter(field_type=FieldType.FLOAT), return_count_filters=True)
    filter_string.update(float_string)

    if request.GET.get('search'):
        favorite.search = request.GET.get('search')
        favorite.updated_at = timezone.now()
        favorite.save()
    
    if request.GET.get('order_by'):
        favorite.order_by = request.GET.get('order_by')
        favorite.updated_at = timezone.now()
        favorite.save()

    base_queryset = favorite_model.objects.filter(combined_q_objects).filter(**filter_string).filter(**snapshot_filter).filter(**pre_filters)
    order_by = request.GET.get('order_by', 'pk')
    queryset = base_queryset
    if hasattr(favorite_model, 'parent'):
        queryset = queryset.filter(parent=None)
        try:
            queryset = queryset.filter(action_status=ActionStatus.IS_ACTIVE)
        except:
            pass
    
    if user_query_conditions:
        queryset = queryset.filter(user_query_conditions)
    if date_query_conditions:
        queryset = queryset.filter(date_query_conditions)
    if int_query_conditions:
        queryset = queryset.filter(int_query_conditions)
    if float_query_conditions:
        queryset = queryset.filter(float_query_conditions)

    if user_count_filters:
        queryset = common_count_filter(user_count_filters, base_queryset, queryset, db_field_names)
    elif date_count_filters:
        queryset = common_count_filter(date_count_filters, base_queryset, queryset, db_field_names)
    elif int_count_filters:
        queryset = common_count_filter(int_count_filters, base_queryset, queryset, db_field_names)
    elif float_count_filters:
        queryset = common_count_filter(float_count_filters, base_queryset, queryset, db_field_names)
    else:
        if order_by in ['count', '-count']:
            order_by = 'pk'

    queryset = queryset.order_by(order_by)

    table_name = f"{favorite.content_type.app_label}_{favorite.content_type.model}"
    if user_unique_filter:
        queryset = common_unique_filter(request, user_unique_filter, queryset, snapshot_filter, table_name)
    if date_unique_filter:
        queryset = common_unique_filter(request, date_unique_filter, queryset, snapshot_filter, table_name)
    if int_unique_filter:
        queryset = common_unique_filter(request, int_unique_filter, queryset, snapshot_filter, table_name)
    if float_unique_filter:
        queryset = common_unique_filter(request, float_unique_filter, queryset, snapshot_filter, table_name)

    favorite_list = software_filter(request, queryset, db_field_names)

    page = request.GET.get('page', 1)
    paginator = Paginator(favorite_list, items)
    
    favorite_items = paginator.page(page)

    pre_column = favorite.pre_columns
    richtext_fields = favorite.richtext_fields if favorite.richtext_fields else []

    total_items = favorite_model.objects.filter(**snapshot_filter).count()
    try:
        total_items = favorite_model.objects.filter(parent=None).filter(action_status=ActionStatus.IS_ACTIVE).filter(**snapshot_filter).count()
    except:
        pass

    fields = get_model_fields(f'{favorite_model.__name__}', pre_column)
    saved_filters = list(SavedFilter.objects.filter(
        userID=request.user.id, parent=ModelChoices.FAVORITE, favorite_id=favorite.id
    ).values())
    for filter in saved_filters:
        if 'created_at' in filter:
            filter['created_at'] = filter['created_at'].isoformat()
    saved_filters_json = json.dumps(saved_filters)

    context = {
        'favorite': favorite,
        'db_field_names': db_field_names,
        'favorite_items': favorite_items,
        'pre_column': pre_column,
        'richtext_fields': richtext_fields,
        'total_items': total_items,
        'date_picker_fields': favorite_model.date_fields_to_convert if favorite_model.date_fields_to_convert else [],
        'int_fields': favorite_model.integer_fields if favorite_model.integer_fields else [],
        'float_fields': favorite_model.float_fields if favorite_model.float_fields else [],
        'snapshots': snapshots,
        'selected_snapshot': summary,
        'latest_snapshot': latest_snapshot,
        'is_dynamic_query': favorite.is_dynamic_query,
        'query_snapshot': query_snapshot if favorite.is_dynamic_query else '',
        'fields': fields,
        'saved_filters': saved_filters_json,
    }
    return render(request, 'apps/favorite/favorite_details.html', context)


def remove_favorite(request, id):
    if request.method == 'POST':
        favorite = Favorite.objects.get(id=id)
        favorite.delete()

        return redirect('/index')


def add_favorite_description(request, favorite_id):
    favorite = Favorite.objects.get(id=favorite_id)
    if request.method == 'POST':
        description = request.POST.get('description')
        favorite.description = description
        favorite.updated_at = timezone.now()
        favorite.save()
        
    return redirect(request.META.get('HTTP_REFERER')) 



def export_favorite_csv_view(request, id):
    favorite = Favorite.objects.get(id=id)
    content_type = ContentType.objects.get(app_label=favorite.content_type.app_label, model=favorite.content_type.model)
    db_field_names = [field.name for field in content_type.model_class()._meta.get_fields() if not field.is_relation and not '_Q_' in field.name]
    favorite_model = content_type.model_class()

    # Snapshot
    latest_snapshot = ''
    summary = None
    snapshot_filter = {}
    try:
        content_type = ContentType.objects.get(model=favorite_model.__name__.lower())
        snapshots = InstantUpload.objects.filter(content_type=content_type).order_by('-created_at')
        snapshot = request.GET.get('snapshot')

        if snapshot and not snapshot == 'all':
            summary = InstantUpload.objects.get(id=snapshot)
            snapshot_filter['loader_instance'] = summary.pk
        
        elif snapshot and snapshot == 'all':
            snapshot_filter= {}
        
        elif favorite.snapshot == 'latest':
            latest_snapshot = snapshots.latest('created_at')
            snapshot_filter['loader_instance'] = latest_snapshot.pk
        
        elif favorite.snapshot and not favorite.snapshot == 'all':
            summary = InstantUpload.objects.get(id=int(favorite.snapshot))
            snapshot_filter['loader_instance'] = int(favorite.snapshot)
        
        elif favorite.snapshot and favorite.snapshot == 'all':
            snapshot_filter= {}
        

    except:
        pass

    fields = []
    pre_column = favorite.pre_columns
    richtext_fields = favorite.richtext_fields if favorite.richtext_fields else []
    show_fields = favorite.hide_show_filters.all().filter(value=False)
    for field in show_fields:
        if field.key not in pre_column:
            fields.append(field.key)

    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{favorite_model._meta.model_name}.csv"'

    
    filter_string = {}
    combined_q_objects, user_unique_filter, user_query_conditions, user_count_filters = same_key_filter(favorite.saved_filters.filter(field_type=FieldType.TEXT), return_count_filters=True)

    pre_filters = {}
    if favorite.pre_filters:
        pre_filters = eval(favorite.pre_filters)

    # for date range
    date_string, date_unique_filter, date_query_conditions, date_count_filters = common_date_filter(favorite.saved_filters.filter(field_type=FieldType.DATE), return_count_filters=True)
    filter_string.update(date_string)
    
    # for integer range
    int_string, int_unique_filter, int_query_conditions, int_count_filters = common_integer_filter(favorite.saved_filters.filter(field_type=FieldType.INT), return_count_filters=True)
    filter_string.update(int_string)

    # for float range
    float_string, float_unique_filter, float_query_conditions, float_count_filters = common_float_filter(favorite.saved_filters.filter(field_type=FieldType.FLOAT), return_count_filters=True)
    filter_string.update(float_string)

    base_queryset = favorite_model.objects.filter(**filter_string).filter(**snapshot_filter).filter(combined_q_objects).filter(**pre_filters)
    order_by = request.GET.get('order_by', 'pk')
    queryset = base_queryset

    if user_query_conditions:
        queryset = queryset.filter(user_query_conditions)
    if date_query_conditions:
        queryset = queryset.filter(date_query_conditions)
    if int_query_conditions:
        queryset = queryset.filter(int_query_conditions)
    if float_query_conditions:
        queryset = queryset.filter(float_query_conditions)

    if user_count_filters:
        queryset = common_count_filter(user_count_filters, base_queryset, queryset, fields)
    elif date_count_filters:
        queryset = common_count_filter(date_count_filters, base_queryset, queryset, fields)
    elif int_count_filters:
        queryset = common_count_filter(int_count_filters, base_queryset, queryset, fields)
    elif float_count_filters:
        queryset = common_count_filter(float_count_filters, base_queryset, queryset, fields)
    else:
        if order_by in ['count', '-count']:
            order_by = 'pk'
        if 'count' in fields:
            fields.remove('count')

    queryset = queryset.order_by(order_by)

    table_name = f"{favorite.content_type.app_label}_{favorite.content_type.model}"
    if user_unique_filter:
        queryset = common_unique_filter(request, user_unique_filter, queryset, snapshot_filter, table_name)
    if date_unique_filter:
        queryset = common_unique_filter(request, date_unique_filter, queryset, snapshot_filter, table_name)
    if int_unique_filter:
        queryset = common_unique_filter(request, int_unique_filter, queryset, snapshot_filter, table_name)
    if float_unique_filter:
        queryset = common_unique_filter(request, float_unique_filter, queryset, snapshot_filter, table_name)

    favorite_list = software_filter(request, queryset, db_field_names)

    writer = csv.writer(response)
    writer.writerow(fields) 

    for fav in favorite_list:
        row_data = []
        for field in fields:
            if field in favorite_model.date_fields_to_convert:
                unix_timestamp = getattr(fav, field)
                if unix_timestamp:
                    date_time = datetime.utcfromtimestamp(unix_timestamp)
                    formatted_date = date_time.strftime('%b %d %Y %I:%M%p')
                    row_data.append(formatted_date)

            elif field in richtext_fields:
                attribute = getattr(fav, field, None)
                if attribute and hasattr(attribute, 'html'):
                    row_data.append(attribute.html)
                else:
                    row_data.append("")
            else:
                row_data.append(getattr(fav, field))
        writer.writerow(row_data)

    return response

def check_task_status(request, task_id):
    try:
        task_status = TaskStatus.objects.get(task_id=task_id)
        return JsonResponse({'is_completed': task_status.is_completed})
    except TaskStatus.DoesNotExist:
        return JsonResponse({'is_completed': False})



#############
# File Upload
#############
from apps.common.forms import FileForm
from apps.common.models import File, FileStatus

@login_required(login_url='/users/signin/')
def file_upload(request):
    files = File.objects.filter(status=FileStatus.ACTIVE)
    form = FileForm()
    if request.method == 'POST':
        form = FileForm(request.POST, request.FILES)
        if form.is_valid():
            files = request.FILES.getlist('files')
            for file in files:
                File.objects.create(
                    description=form.cleaned_data.get('description'),
                    file=file,
                    created_by=request.user
                )
            
            return redirect(request.META.get('HTTP_REFERER'))

    context = {
        'files': files,
        'form': form,
        'segment': 'files'
    }

    return render(request, 'pages/file-upload.html', context)


@login_required(login_url='/users/signin/')
def download_file(request, file_id):
    file_instance = get_object_or_404(File, id=file_id)

    if not file_instance.file:
        raise Http404("File not found.")

    filename = os.path.basename(file_instance.file.name)
    filename = f"{file_instance.description}-{filename}"

    response = FileResponse(file_instance.file.open('rb'), as_attachment=True, filename=filename)
    return response


@login_required(login_url='/users/signin/')
def delete_file(request, file_id):
    file_instance = get_object_or_404(File, id=file_id)
    file_instance.status = FileStatus.DELETED
    file_instance.updated_by = request.user
    file_instance.save()
    return redirect(request.META.get('HTTP_REFERER'))


# Global views
import io
from home.models import IPE
from django.utils.timezone import now
from urllib.parse import urlparse
from apps.tables.models import SelectedRows

def delete_saved_filter(request):
    if request.method == 'POST':
        filter_id = request.POST.get('filter_id')
        try:
            # Delete the filter from the database
            filter_to_delete = SavedFilter.objects.get(id=filter_id, userID=request.user.id)
            filter_to_delete.delete()
            return JsonResponse({'success': True})
        except SavedFilter.DoesNotExist:
            return JsonResponse({'success': False, 'error': 'Filter not found'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})


import pandas as pd


def loader_csv_upload(request, model_name):
    try:
        content_type = ContentType.objects.get(app_label="tables", model=model_name)
    except ContentType.DoesNotExist:
        return redirect(request.META.get('HTTP_REFERER'))
    
    if request.method == 'POST':
        uploaded_file = request.FILES.get('csv_file')
        if uploaded_file:
            file_name = uploaded_file.name
            file_extension = file_name.split('.')[-1].lower()

            if file_extension in ['xls', 'xlsx']:
                df = pd.read_excel(uploaded_file, dtype=str, engine='openpyxl')
                df = df.fillna('')  
                column_names = ", ".join(df.columns) if not df.empty else "No columns detected"
                row_count = len(df)
            else:
                text_file = io.TextIOWrapper(uploaded_file.file, encoding='latin-1', errors='replace')
                reader = csv.reader(text_file)

                columns = next(reader, []) 
                column_names = ", ".join(columns) if columns else "No columns detected"
                row_count = sum(1 for _ in reader)
                text_file.seek(0)

            snapshot = InstantUpload.objects.create(
                type=TypeChoices.FILE,
                content_type=content_type,
                file=uploaded_file
            )

            user = request.user
            timestamp_utc = timezone.now()
            timestamp_local = timezone.localtime(timestamp_utc)
            timestamp_str = timestamp_local.strftime('%Y-%m-%d %H:%M:%S')

            IPE.objects.create(
                userID=user,
                path=urlparse(request.META.get('HTTP_REFERER', request.path)).path,
                description=f'Uploaded by: {user.get_full_name() or user.username}\n'
                            f'Uploaded on: {timestamp_str}\n'
                            f'File name: {file_name}\n'
                            f'Rows extracted: {row_count}\n'
                            f'Snapshot: {snapshot.pk}\n'
                            f'Columns: {column_names}'
            )

    return redirect(request.META.get('HTTP_REFERER'))

def ipe_details(request):
    path = request.GET.get('path')
    domain = request.build_absolute_uri('/')[:-1]
    full_url = f"{domain}{path}"
    ipe_qs = IPE.objects.filter(path=path).order_by('-created_at')
    context = {
        'ipe_qs': ipe_qs,
        'full_url': full_url
    }
    return render(request, 'pages/ipe-details.html', context)


def save_selected_rows(request):
    model = request.GET.get('model')
    model_choice = request.GET.get('model_choice')
    ids = request.GET.get('ids')
    tab_id = request.GET.get('tab_id')
    finding_id = request.GET.get('finding_id')
    unique_id = request.GET.get('unique_id')
    SelectedRows.objects.filter(
        model=model, model_choice=model_choice,
        tab_id=tab_id, finding_id=finding_id, unique_id=unique_id
    ).delete()
    if model and model_choice and ids:
        SelectedRows.objects.create(
            model=model, model_choice=model_choice, rows=ids,
            tab_id=tab_id, finding_id=finding_id, unique_id=unique_id
        )
    return redirect(request.META.get('HTTP_REFERER'))

def delete_unmatched(request):
    model = request.GET.get('model')
    model_choice = request.GET.get('model_choice')
    ipe_path = request.GET.get('ipe_path')

    last_ipe = IPE.objects.filter(path=ipe_path).last()
    base_description = last_ipe.description if last_ipe else ""

    user = request.user

    ids = request.GET.get('ids')
    if model and model_choice and ids:
        model_class = apps.get_model('tables', model)
        join_model = get_object_or_404(JoinModel, pk=model_class.join_model_instance)
        ids = ids.split(',')
        reduced_rows = model_class.objects.filter(pk__in=ids)

        base_fields = [field.strip() for field in join_model.base_f_key_fields.split(',')] if join_model.base_f_key_fields else []
        delta_fields = [f"d_{field.strip()}" for field in join_model.delta_f_key_fields.split(',')] if join_model.delta_f_key_fields else []

        valid_field_names = {field.name for field in model_class._meta.get_fields()}

        requested_fields = base_fields + delta_fields
        fields_to_use = [f for f in requested_fields if f in valid_field_names]

        if not fields_to_use:
            fields_to_use = ['id']

        new_description = (
            f"{base_description}\n" if base_description else ""
        ) + f"Reduced rows:\n" + "\n".join([str(row) for row in reduced_rows.values(*fields_to_use)])


        IPE.objects.create(
            userID=user,
            path=ipe_path,
            description=new_description
        )

        reduced_rows.delete()

    return redirect(request.META.get('HTTP_REFERER'))


def create_py_func_prompt(request):
    PyFunctionPrompt.objects.all().delete()
    if request.method == 'POST':
        prompt = request.POST.get('py_prompt')
        PyFunctionPrompt.objects.create(prompt=prompt)
        return redirect(request.META.get('HTTP_REFERER'))
    
    return redirect(request.META.get('HTTP_REFERER'))


def create_py_func(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        func = request.POST.get('function')
        description = request.POST.get('description')
        PyFunction.objects.update_or_create(
            name=name,
            defaults={
                'func': func,
                'description': description
            }
        )
        return redirect(request.META.get('HTTP_REFERER'))
    
    return redirect(request.META.get('HTTP_REFERER'))


def edit_py_func(request):
    if request.method == 'POST':
        func_id = request.POST.get('func_id')
        py_func = get_object_or_404(PyFunction, id=func_id)

        name = request.POST.get('name')
        func = request.POST.get('function')
        description = request.POST.get('description')

        if not PyFunction.objects.exclude(id=py_func.id).filter(name=name).exists():
            py_func.name = name
            py_func.func = func
            py_func.description = description
            py_func.save()
        else:
            print("Function with this name already exists")

        return redirect(request.META.get('HTTP_REFERER'))

    return redirect(request.META.get('HTTP_REFERER'))


# Stored Procedure
from django.apps import apps
from loader.models import database_connection, check_valid_unix, StoredProcedure, hash_json, json_to_vector, is_pgvector_enabled
from celery import shared_task
import base64

def query_shoot_func(model, procedure, loader_instance_pk):
    result = []
    
    if procedure.database:
        result, row_count, columns = database_connection(procedure.database, procedure.procedure)
    else:
        with connection.cursor() as cursor:
            cursor.execute(procedure.procedure)
            result = cursor.fetchall()
            columns = [col[0] for col in cursor.description]

    instances_to_create = []
    db_field_names = [field.name for field in model._meta.get_fields() if not field.is_relation]

    db_field_names = [field for field in db_field_names if field.lower() != "id"]
    id_index = next((i for i, col in enumerate(columns) if col.lower() == "id"), None)
    if id_index is not None:
        columns.pop(id_index)

    date_fields_to_convert = getattr(model, 'date_fields_to_convert', [])
    ad_unix_dates = getattr(model, 'ad_unix_dates', [])
    unix_dates = getattr(model, 'unix_dates', [])

    possible_formats = [
        '%b %d %Y %I:%M%p', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d',
        '%d/%m/%Y %H:%M', '%Y/%m/%d %H:%M', '%Y%m%d%H%M%S.%fZ', '%b %d, %Y %I:%M %p',
        '%d/%m/%Y %I:%M %p', '%d/%m/%Y', '%d %b, %Y', '%m/%d/%Y, %H:%M:%S',
        '%m/%d/%Y, %I:%M:%S %p', '%m/%d/%Y %I:%M:%S %p', '%Y/%m/%d', 
        '%m/%d/%Y %I:%M:%S %p', '%Y/%m/%d %H:%M:%S', '%d/%m/%y', '%d %B %Y %H:%M', 
        '%Y/%m/%d', '%d-%b-%Y', '%Y-%m-%d %H:%M:%S.%f'
    ]

    for row in result:
        row = list(row)
        if id_index is not None:
            row.pop(id_index)

        field_values = dict(zip(db_field_names, row))
        field_values["loader_instance"] = loader_instance_pk

        # **Apply Date Conversion**
        for column_name in date_fields_to_convert:
            if column_name in field_values:
                value = field_values[column_name]
                
                if value is None or value == "":
                    continue
                
                try:
                    if isinstance(value, (int, float)) or (isinstance(value, str) and value.isdigit()):
                        value_int = int(value)
                        
                        if column_name in ad_unix_dates and value_int != 0:
                            unix_timestamp = (value_int / 10000000) - 11644473600
                            field_values[column_name] = int(unix_timestamp)
                        elif column_name in unix_dates:
                            unix_timestamp = datetime.fromtimestamp(value_int, tz=timezone.utc)
                            field_values[column_name] = int(unix_timestamp.timestamp())
                    
                    else:
                        value = str(value).strip()
                        date_object = None

                        for format_str in possible_formats:
                            try:
                                date_object = datetime.strptime(value, format_str)
                                break
                            except ValueError:
                                pass
                        
                        if date_object:
                            unix_timestamp = check_valid_unix(date_object)
                            field_values[column_name] = unix_timestamp

                except ValueError:
                    pass
            
        
        db_alias = model.objects.db
        database_path = connections[db_alias].settings_dict['NAME']
        database_name = os.path.basename(database_path) if database_path else database_path
        json_data = {
            "database": database_name,
            "table": model._meta.db_table,
            "snapshot_at": now().isoformat(),
        }
        for key, value in field_values.items():
            if isinstance(value, bytes):
                json_data[key] = base64.urlsafe_b64encode(value).decode()
            else:
                json_data[key] = value

        if is_pgvector_enabled():
            hash_data = json_to_vector(json_data)
        else:
            hash_data = hash_json(json_data)
        if 'json_data' in [field.name for field in model._meta.get_fields()]:
            field_values['json_data'] = json_data

        if 'hash_data' in [field.name for field in model._meta.get_fields()]:
            field_values['hash_data'] = hash_data

        instances_to_create.append(model(**field_values))

    return instances_to_create


@shared_task
def fetch_stored_procedure_data(user_id, referer, name, model_name):
    print("Start Fetch")
    procedure = get_object_or_404(StoredProcedure, name=name)
    print("procedure")
    model = apps.get_model('tables', model_name)
    content_type = ContentType.objects.get_for_model(model)
    model_class = content_type.model_class()
    print("model content_type model_class")
    loader_instance = InstantUpload.objects.create(
        type=TypeChoices.SP,
        content_type=content_type
    )
    print("instances_to_create")
    instances_to_create = query_shoot_func(model_class, procedure, loader_instance.pk)
    row_count = len(instances_to_create) 
    model_class.objects.bulk_create(instances_to_create)

    user = User.objects.get(id=user_id)
    timestamp_utc = timezone.now()
    timestamp_local = timezone.localtime(timestamp_utc)
    timestamp_str = timestamp_local.strftime('%Y-%m-%d %H:%M:%S')

    column_names = [field.name for field in model_class._meta.get_fields() if not field.is_relation]  

    IPE.objects.create(
        userID=user,
        path=urlparse(referer).path,
        description=f'SP Executed by: {user.get_full_name() or user.username}\n'
                    f'SP Executed on: {timestamp_str}\n'
                    f'Rows extracted: {row_count}\n'
                    f'Snapshot: {loader_instance.pk}\n'
                    f'Columns: {", ".join(column_names)}'
    )

    print("end")

def trigger_fetch_stored_procedure_data(request, name, model_name):
    user_id = request.user.id
    referer = request.META.get('HTTP_REFERER', request.path)

    fetch_stored_procedure_data.delay(user_id, referer, name, model_name)
    return redirect(referer)



def loader_query_upload(request, model_name):
    try:
        content_type = ContentType.objects.get(app_label="tables", model=model_name)
    except ContentType.DoesNotExist:
        return redirect(request.META.get('HTTP_REFERER'))
    
    if request.method == 'POST':
        query_pk = request.POST.get('query')
        if query_pk:
            query = get_object_or_404(DynamicQuery, pk=query_pk)

            snapshot = InstantUpload.objects.create(
                type=TypeChoices.QUERY,
                content_type=content_type,
                query=query
            )

            with connection.cursor() as cursor:
                cursor.execute(query.query)
                result = cursor.fetchall()
                column_names = [col[0] for col in cursor.description] if cursor.description else []
                row_count = len(result)

            user = request.user

            timestamp_utc = timezone.now()
            timestamp_local = timezone.localtime(timestamp_utc)
            timestamp_str = timestamp_local.strftime('%Y-%m-%d %H:%M:%S')

            IPE.objects.create(
                userID=user,
                path=urlparse(request.META.get('HTTP_REFERER', request.path)).path,
                description=f'Query Executed by: {user.get_full_name() or user.username}\n'
                            f'Query Executed on: {timestamp_str}\n'
                            f'SQL Query: {query.query}\n'
                            f'Rows extracted: {row_count}\n'
                            f'Snapshot: {snapshot.pk}\n'
                            f'Columns: {column_names}'
            )
        return redirect(request.META.get('HTTP_REFERER'))


def sap_api_shoot_func(model, api_instance, loader_instance_pk):
    result = []

    headers = {
        'APIKey': api_instance.api_key,
        'Accept': 'application/json',
        'DataServiceVersion': '2.0',
    }

    params = {
        '$top': api_instance.top
    }

    response = requests.get(api_instance.api_url, headers=headers, params=params)

    if response.status_code != 200:
        raise Exception(f"Failed to fetch from SAP API. Status {response.status_code}: {response.text}")

    data = response.json()
    if 'value' not in data or not isinstance(data['value'], list):
        return []

    result = data['value']
    if not result:
        return []

    instances_to_create = []
    db_field_names = [field.name for field in model._meta.get_fields() if not field.is_relation]
    db_field_names = [field for field in db_field_names if field.lower() != "id"]

    date_fields_to_convert = getattr(model, 'date_fields_to_convert', [])
    ad_unix_dates = getattr(model, 'ad_unix_dates', [])
    unix_dates = getattr(model, 'unix_dates', [])

    possible_formats = [
        '%b %d %Y %I:%M%p', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d',
        '%d/%m/%Y %H:%M', '%Y/%m/%d %H:%M', '%Y%m%d%H%M%S.%fZ', '%b %d, %Y %I:%M %p',
        '%d/%m/%Y %I:%M %p', '%d/%m/%Y', '%d %b, %Y', '%m/%d/%Y, %H:%M:%S',
        '%m/%d/%Y, %I:%M:%S %p', '%m/%d/%Y %I:%M:%S %p', '%Y/%m/%d', 
        '%m/%d/%Y %I:%M:%S %p', '%Y/%m/%d %H:%M:%S', '%d/%m/%y', '%d %B %Y %H:%M', 
        '%Y/%m/%d', '%d-%b-%Y', '%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%dT%H:%M:%SZ'
    ]

    for row_data in result:
        field_values = {k: v for k, v in row_data.items() if k in db_field_names}
        field_values["loader_instance"] = loader_instance_pk

        # Convert Dates
        for column_name in date_fields_to_convert:
            if column_name in field_values:
                value = field_values[column_name]
                if value is None or value == "":
                    continue
                try:
                    if isinstance(value, (int, float)) or (isinstance(value, str) and value.isdigit()):
                        value_int = int(value)

                        if column_name in ad_unix_dates and value_int != 0:
                            unix_timestamp = (value_int / 10000000) - 11644473600
                            field_values[column_name] = int(unix_timestamp)
                        elif column_name in unix_dates:
                            unix_timestamp = datetime.fromtimestamp(value_int, tz=timezone.utc)
                            field_values[column_name] = int(unix_timestamp.timestamp())
                    else:
                        value = str(value).strip()
                        date_object = None
                        for format_str in possible_formats:
                            try:
                                date_object = datetime.strptime(value, format_str)
                                break
                            except ValueError:
                                continue
                        if date_object:
                            unix_timestamp = check_valid_unix(date_object)
                            field_values[column_name] = unix_timestamp
                except ValueError:
                    continue

        db_alias = model.objects.db
        database_path = connections[db_alias].settings_dict['NAME']
        database_name = os.path.basename(database_path) if database_path else database_path

        json_data = {
            "database": database_name,
            "table": model._meta.db_table,
            "snapshot_at": now().isoformat(),
        }

        for key, value in field_values.items():
            if isinstance(value, bytes):
                json_data[key] = base64.urlsafe_b64encode(value).decode()
            else:
                json_data[key] = value

        if is_pgvector_enabled():
            hash_data = json_to_vector(json_data)
        else:
            hash_data = hash_json(json_data)

        if 'json_data' in [field.name for field in model._meta.get_fields()]:
            field_values['json_data'] = json_data

        if 'hash_data' in [field.name for field in model._meta.get_fields()]:
            field_values['hash_data'] = hash_data

        instances_to_create.append(model(**field_values))

    return instances_to_create



@shared_task
def fetch_api_data(user_id, referer, id, model_name):
    print("Start Fetch")
    sap_api = get_object_or_404(SAPApi, id=id)
    model = apps.get_model('tables', model_name)
    content_type = ContentType.objects.get_for_model(model)
    model_class = content_type.model_class()
    print("model content_type model_class")
    loader_instance = InstantUpload.objects.create(
        type=TypeChoices.SP,
        content_type=content_type
    )
    print("instances_to_create")
    instances_to_create = sap_api_shoot_func(model_class, sap_api, loader_instance.pk)
    row_count = len(instances_to_create) 
    model_class.objects.bulk_create(instances_to_create)

    user = User.objects.get(id=user_id)
    timestamp_utc = timezone.now()
    timestamp_local = timezone.localtime(timestamp_utc)
    timestamp_str = timestamp_local.strftime('%Y-%m-%d %H:%M:%S')

    column_names = [field.name for field in model_class._meta.get_fields() if not field.is_relation]  

    IPE.objects.create(
        userID=user,
        path=urlparse(referer).path,
        description=f'SAP API Run by: {user.get_full_name() or user.username}\n'
                    f'SAP API Run on: {timestamp_str}\n'
                    f'SAP API URL: {sap_api.api_url}\n'
                    f'Rows extracted: {row_count}\n'
                    f'Snapshot: {loader_instance.pk}\n'
                    f'Columns: {", ".join(column_names)}'
    )

    print("end")

def trigger_fetch_api_data(request, model_name):
    if request.method == 'POST':
        sap_id = request.POST.get('sap_api')
        user_id = request.user.id
        referer = request.META.get('HTTP_REFERER', request.path)

        fetch_api_data.delay(user_id, referer, sap_id, model_name)
        return redirect(referer)
    return redirect(request.META.get('HTTP_REFERER'))


def export_ipe_description(request, ipe_id):
    ipe = get_object_or_404(IPE, id=ipe_id)
    path = ipe.path
    tab = False
    tab_id = None
    page_name = None

    if '/tab/' in path:
        match = re.search(r'/tables/tab/(.+)/', path)
        if not match:
            raise Http404("Invalid path format in IPE")
        tab_id = match.group(1)
        url_name = "export_tab_excel_view"
        tab = True
    else:
        match = re.search(r'/tables/(.+)/', path)
        if not match:
            raise Http404("Invalid path format in IPE")
        page_name = match.group(1)
        url_name = f"export_{page_name}_excel"

    wb = Workbook()
    ws = wb.active
    ws.title = "Export Data"

    title_font = Font(bold=True, size=14, color='FFFFFF')
    header_font = Font(bold=True)
    fill = PatternFill(start_color="033d6d", end_color="033d6d", fill_type="solid")
    title_column_span = 25

    _add_title_section(ws, "IPE", title_font, fill, title_column_span)

    description_dict = ipe.get_description_dict()
    for key, value in description_dict.items():
        ws.append([key, value])
        ws.cell(row=ws.max_row, column=1).font = header_font

    try:
        if tab:
            export_wb = _get_tab_export_workbook(request, url_name, tab_id)
        else:
            export_wb = _get_standard_export_workbook(request, url_name)

        export_ws = export_wb.active

        start_row = ws.max_row + 3
        _add_title_section(ws, "Evidence", title_font, fill, title_column_span, start_row)

        header_row = start_row + 1
        for col, cell in enumerate(export_ws[1], start=1):
            ws.cell(row=header_row, column=col, value=cell.value).font = header_font

        for row_data in export_ws.iter_rows(min_row=2, values_only=True):
            header_row += 1
            for col, value in enumerate(row_data, start=1):
                ws.cell(row=header_row, column=col, value=value)

    except Exception as e:
        return HttpResponse(
            f"Error processing export data: {str(e)}<br>URL Name: {url_name}",
            status=500
        )

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"tab_with_ipe.xlsx" if tab else f"{page_name}_with_ipe.xlsx"
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename={filename}'
    return response


def _add_title_section(ws, title, font, fill, column_span, start_row=1):
    ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row, end_column=column_span)
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = font
    cell.fill = fill
    ws.row_dimensions[start_row].height = 25
    for col in range(1, column_span + 1):
        ws.cell(row=start_row, column=col).fill = fill


def _get_tab_export_workbook(request, url_name, tab_id):
    export_url = reverse(url_name, kwargs={'id': str(tab_id)})
    resolver_match = resolve(export_url)

    if hasattr(resolver_match.func, 'view_class'):
        export_view = resolver_match.func.view_class()
        workbook = export_view.get_excel_workbook(request, tab_id)
        return workbook 
    else:
        response = resolver_match.func(request, tab_id)
        return response


def _get_standard_export_workbook(request, url_name):
    export_url = reverse(url_name)
    resolver_match = resolve(export_url)

    if hasattr(resolver_match.func, 'view_class'):
        export_view = resolver_match.func.view_class()
        return export_view.get_excel_workbook(request)
    else:
        return resolver_match.func(request)


@csrf_exempt
def save_column_order(request):
    if request.method == 'POST':
        data = json.loads(request.body)
        table_name = data.get('table_name')
        column_order = data.get('column_order')

        favorite_id = data.get('favorite_id')
        finding_id = data.get('finding_id')
        img_loader_id = data.get('img_loader_id')
        unique_id = data.get('unique_id')
        tab_id = data.get('tab_id')

        filter_kwargs = {
            'user': request.user,
            'table_name': table_name
        }

        if favorite_id:
            filter_kwargs['favorite_id'] = favorite_id
            filter_kwargs['finding_id__isnull'] = True
            filter_kwargs['img_loader_id__isnull'] = True
            filter_kwargs['unique_id__isnull'] = True
            filter_kwargs['tab_id__isnull'] = True
        elif finding_id:
            filter_kwargs['favorite_id__isnull'] = True
            filter_kwargs['finding_id'] = finding_id
            filter_kwargs['img_loader_id__isnull'] = True
            filter_kwargs['unique_id__isnull'] = True
            filter_kwargs['tab_id__isnull'] = True
        elif img_loader_id:
            filter_kwargs['favorite_id__isnull'] = True
            filter_kwargs['finding_id__isnull'] = True
            filter_kwargs['img_loader_id'] = img_loader_id
            filter_kwargs['unique_id__isnull'] = True
            filter_kwargs['tab_id__isnull'] = True
        elif unique_id:
            filter_kwargs['favorite_id__isnull'] = True
            filter_kwargs['finding_id__isnull'] = True
            filter_kwargs['img_loader_id__isnull'] = True
            filter_kwargs['unique_id'] = unique_id
            filter_kwargs['tab_id__isnull'] = True
        elif tab_id:
            filter_kwargs['favorite_id__isnull'] = True
            filter_kwargs['finding_id__isnull'] = True
            filter_kwargs['img_loader_id__isnull'] = True
            filter_kwargs['unique_id__isnull'] = True
            filter_kwargs['tab_id'] = tab_id
        else:
            filter_kwargs['favorite_id__isnull'] = True
            filter_kwargs['finding_id__isnull'] = True
            filter_kwargs['img_loader_id__isnull'] = True
            filter_kwargs['unique_id__isnull'] = True
            filter_kwargs['tab_id__isnull'] = True

        defaults = {
            'column_order': column_order,
            'favorite_id': favorite_id,
            'finding_id': finding_id,
            'img_loader_id': img_loader_id,
            'unique_id': unique_id,
            'tab_id': tab_id,
        }

        obj, created = ColumnOrder.objects.update_or_create(
            defaults=defaults,
            **filter_kwargs
        )
        
        return JsonResponse({'status': 'success'})

    return JsonResponse({'error': 'Invalid method'}, status=400)