import csv
import json
import openpyxl
from apps.common.models import SavedFilter, FieldType
from django.http import HttpResponse
from django.shortcuts import render, redirect
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.contrib.auth.decorators import login_required
from apps.tables.utils import (
    software_filter, 
    same_key_filter, 
    common_unique_filter,
    common_count_filter,
    get_user_id,
    common_snapshot_filter, get_model_fields,
    common_date_filter, common_integer_filter, common_float_filter
)
from apps.tables.models import (
    UserFilter, 
    PageItems, 
    HideShowFilter, 
    DateRangeFilter, 
    IntRangeFilter, 
    FloatRangeFilter, 
    ModelChoices, 
    #DeviceView,
    Finding
)
from django.http import HttpResponse
from django.views import View
from django.http import JsonResponse
from datetime import datetime
from django.db.models import F
from django.apps import apps
from django.urls import reverse
from urllib.parse import unquote
from openpyxl.utils import get_column_letter

from apps.tables.calculations import calculate_contract_value, calculate_remaining_days
    

# Create your views here.

def create_unique_details_filter(request, unique_model_name, unique_field_name):
    unique_id = f"{unique_model_name}.{unique_field_name}"
    model = apps.get_model('tables', unique_model_name)

    field_type_map = {
        'date': getattr(model, 'date_fields_to_convert', []),
        'int': getattr(model, 'integer_fields', []),
        'float': getattr(model, 'float_fields', [])
    }

    def get_field_type(value):
        for f_type, fields in field_type_map.items():
            if value in fields:
                return getattr(FieldType, f_type.upper())
        return FieldType.TEXT

    def get_boolean_field(name, index):
        return request.POST.get(f'{name}_{index}') == 'true'

    existing_count_filter = SavedFilter.objects.filter(
        userID=request.user.id, parent=ModelChoices.UNIQUE, is_count=True, unique_id=unique_id
    ).first()
    for key, value in request.POST.items():
        if key.startswith('field_name_') and value:
            index = key.split('_')[-1]
            field_name = value
            field_type = get_field_type(value)

            if field_type == FieldType.TEXT:
                value_start = request.POST.get(f'value_start_{index}') or ''
                value_end = request.POST.get(f'value_end_{index}') or ''
            else:
                value_start = request.POST.get(f'value_start_{index}') or None
                value_end = request.POST.get(f'value_end_{index}') or None

            is_null = get_boolean_field('is_null', index)
            is_not_null = get_boolean_field('is_not_null', index)
            is_unique = get_boolean_field('is_unique', index)
            is_count = get_boolean_field('is_count', index)

            filter_id = request.POST.get(f'filter_id_{index}')
            if filter_id in [None, '', 'null']:
                filter_id = None
            
            if is_count:
                if existing_count_filter and (not filter_id or str(existing_count_filter.id) != filter_id):
                    is_count = False

            filter_data = {
                'userID': request.user.id,
                'parent': ModelChoices.UNIQUE,
                'field_name': field_name,
                'field_type': field_type,
                'value_start': value_start,
                'value_end': value_end,
                'is_null': is_null,
                'is_not_null': is_not_null,
                'is_unique': is_unique,
                'is_count': is_count,
                'unique_id': unique_id
            }

            if filter_id:
                SavedFilter.objects.filter(pk=filter_id).update(**filter_data)
            else:
                SavedFilter.objects.create(**filter_data)
    
    return redirect(request.META.get('HTTP_REFERER'))

def create_unique_details_page_items(request, unique_id):
    if request.method == 'POST':
        items = request.POST.get('items')
        PageItems.objects.update_or_create(
            userID=get_user_id(request),
            parent=ModelChoices.UNIQUE,
            unique_id=unique_id,
            defaults={'items_per_page':items}
        )

        return redirect(reverse('list_of_devices'))

def create_unique_details_hide_show_filter(request, unique_id):
    if request.method == "POST":
        data_list = json.loads(request.body)

        for data in data_list:
            hide_show_filter, created = HideShowFilter.objects.update_or_create(
                userID=get_user_id(request),
                parent=ModelChoices.UNIQUE,
                key=data.get('key'),
                unique_id=unique_id,
                defaults={'value': data.get('value')}
            )

        response_data = {'message': 'Filters updated successfully'}
        return JsonResponse(response_data)

    return JsonResponse({'error': 'Invalid request'}, status=400)


def delete_unique_details_filter(request, id):
    filter_instance = UserFilter.objects.get(id=id, userID=get_user_id(request), parent=ModelChoices.UNIQUE)
    filter_instance.delete()

    return redirect(request.META.get('HTTP_REFERER'))

def delete_unique_details_daterange_filter(request, id):
    filter_instance = DateRangeFilter.objects.get(id=id, userID=get_user_id(request), parent=ModelChoices.UNIQUE)
    filter_instance.delete()

    return redirect(request.META.get('HTTP_REFERER'))

def delete_unique_details_intrange_filter(request, id):
    filter_instance = IntRangeFilter.objects.get(id=id, userID=get_user_id(request), parent=ModelChoices.UNIQUE)
    filter_instance.delete()

    return redirect(request.META.get('HTTP_REFERER'))

def delete_unique_details_floatrange_filter(request, id):
    filter_instance = FloatRangeFilter.objects.get(id=id, userID=get_user_id(request), parent=ModelChoices.UNIQUE)
    filter_instance.delete()

    return redirect(request.META.get('HTTP_REFERER'))


def field_type_mapping(model, field, value):
    if not value:
        return None

    if field in getattr(model, 'date_fields_to_convert', []) or field in getattr(model, 'integer_fields', []):
        return int(value)
    elif field in getattr(model, 'float_fields', []):
        return float(value)
    
    return value

def create_source_filters(request, s_model_choice, unique_id):
    SavedFilter.objects.filter(
        userID=get_user_id(request),
        parent=ModelChoices.UNIQUE,
        unique_id=unique_id,
        s_model_choice=ModelChoices[s_model_choice]
    ).delete()

    source_filters = SavedFilter.objects.filter(
        userID=get_user_id(request),
        parent=ModelChoices[s_model_choice],
        is_unique=False,
        is_count=False
    )

    new_filters = []
    for source_filter in source_filters:
        new_filter = SavedFilter(
            userID=get_user_id(request),
            parent=ModelChoices.UNIQUE,
            field_name=source_filter.field_name,
            field_type=source_filter.field_type,
            value_start=source_filter.value_start,
            value_end=source_filter.value_end,
            is_null=source_filter.is_null,
            is_not_null=source_filter.is_not_null,
            is_unique=source_filter.is_unique,
            is_count=source_filter.is_count,
            s_model_choice=ModelChoices[s_model_choice],
            is_hidden=source_filter.is_hidden,
            favorite_id=source_filter.favorite_id,
            finding_id=source_filter.finding_id,
            img_loader_id=source_filter.img_loader_id,
            unique_id=unique_id,
            tab_id=source_filter.tab_id,
        )
        new_filters.append(new_filter)

    SavedFilter.objects.bulk_create(new_filters)

    print(f"Created {len(new_filters)} new filters")


def unique_details(request, unique_model_name, unique_field_name, unique_value='', finding_id=None, s_model_choice=None):
    unique_value = unquote(unquote(unique_value))
    unique_id = f"{unique_model_name}.{unique_field_name}"
    unique_model = apps.get_model('tables', unique_model_name)
    unique_value = field_type_mapping(unique_model, unique_field_name, unique_value)
    unique_details = unique_model.objects.filter(**{unique_field_name: unique_value})

    label_to_value = {label: value for value, label in ModelChoices.choices}
    s_model_choice = label_to_value.get(unique_model.__name__)

    create_source_filters(request, s_model_choice, unique_id)

    finding = None
    if finding_id:
        finding = Finding.objects.get(id=finding_id)

    db_field_names = [field.name for field in unique_model._meta.get_fields() if not field.is_relation]

    # Snapshot
    latest_snapshot, summary, snapshot_filter, snapshots = common_snapshot_filter(request, unique_model_name)

    field_names = []
    for field_name in db_field_names:
        fields, created = HideShowFilter.objects.get_or_create(key=field_name, userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id)
        field_names.append(fields)

    page_items = PageItems.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id).last()
    items = 25
    if page_items:
        items = page_items.items_per_page

    filter_string = {}
    pre_filter_string = {}
    # pre_filter_string['Surname__icontains'] = ''

    filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.TEXT)
    combined_q_objects, user_unique_filter, user_query_conditions, user_count_filters = same_key_filter(filter_instance, return_count_filters=True)

    # for date range
    date_filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.DATE)
    date_string, date_unique_filter, date_query_conditions, date_count_filters = common_date_filter(date_filter_instance, return_count_filters=True)
    filter_string.update(date_string)
    
    # for integer range
    int_filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.INT)
    int_string, int_unique_filter, int_query_conditions, int_count_filters = common_integer_filter(int_filter_instance, return_count_filters=True)
    filter_string.update(int_string)

    # for float range
    float_filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.FLOAT)
    float_string, float_unique_filter, float_query_conditions, float_count_filters = common_float_filter(float_filter_instance, return_count_filters=True)
    filter_string.update(float_string)

    order_by = request.GET.get('order_by', 'pk')
    base_queryset = unique_details.filter(**filter_string).filter(**pre_filter_string).filter(combined_q_objects).filter(**snapshot_filter)
    if finding and finding.selected_rows:
        base_queryset = base_queryset.filter(pk__in=eval(finding.selected_rows))
    
    if hasattr(unique_model, 'parent'):
        base_queryset = base_queryset.filter(parent=None)

    queryset = base_queryset
    
    if user_query_conditions:
        queryset = queryset.filter(user_query_conditions)
    if date_query_conditions:
        queryset = queryset.filter(date_query_conditions)
    if int_query_conditions:
        queryset = queryset.filter(int_query_conditions)
    if float_query_conditions:
        queryset = queryset.filter(float_query_conditions)
    
    if user_count_filters:
        queryset = common_count_filter(user_count_filters, base_queryset, queryset, db_field_names)
    elif date_count_filters:
        queryset = common_count_filter(date_count_filters, base_queryset, queryset, db_field_names)
    elif int_count_filters:
        queryset = common_count_filter(int_count_filters, base_queryset, queryset, db_field_names)
    elif float_count_filters:
        queryset = common_count_filter(float_count_filters, base_queryset, queryset, db_field_names)
    else:
        if order_by == 'count' or order_by == '-count':
            order_by = 'pk'

    queryset = queryset.order_by(order_by)

    if user_unique_filter:
        queryset = common_unique_filter(request, user_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')
    if date_unique_filter:
        queryset = common_unique_filter(request, date_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')
    if int_unique_filter:
        queryset = common_unique_filter(request, int_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')
    if float_unique_filter:
        queryset = common_unique_filter(request, float_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')

    # Order by
    if order_by == 'count' and 'count' not in db_field_names:
        queryset = queryset.order_by(F('count').desc(nulls_last=True))
    else:
        queryset = queryset.order_by(order_by)

    software_list = software_filter(request, queryset, db_field_names)

    page = request.GET.get('page', 1)
    paginator = Paginator(software_list, items)
    
    try:
        softwares = paginator.page(page)
    except PageNotAnInteger:
        return redirect('unique_details')
    except EmptyPage:
        return redirect('unique_details') 

    read_only_fields = ('ID', 'loader_instance','hash_data', 'json_data',)
    pre_column = ('ID', 'loader_instance','hash_data', 'json_data',)
    compulsory_fields = ()
    
    COMBOS = {}
    #COMBOS['Contract_Type'] = TableDropdownSubItem.objects.filter(item__item='Contract_Type').values_list('subitem', flat=True)

    fields = get_model_fields(f'{unique_model.__name__}', pre_column)
    saved_filters = list(SavedFilter.objects.filter(
        userID=request.user.id, parent=ModelChoices.UNIQUE, unique_id=unique_id
    ).values())
    for filter in saved_filters:
        if 'created_at' in filter:
            filter['created_at'] = filter['created_at'].isoformat()
    saved_filters_json = json.dumps(saved_filters)

    context = {
        'segment'  : 'unique_details',
        'parent'   : 'dashboard',
        'softwares' : softwares,
        'field_names': field_names,
        'db_field_names': db_field_names,
        'filter_instance': filter_instance,
        'date_filter_instance': date_filter_instance,
        'items': items,
        'read_only_fields': read_only_fields,
        'total_items': base_queryset.count(),
        'pre_column': pre_column,
        'COMBOS': COMBOS,
        'date_picker_fields': unique_model.date_fields_to_convert,
        'compulsory_fields': compulsory_fields,
        'fav_name': 'Unique Details',
        'int_filter_instance': int_filter_instance,
        'int_fields': unique_model.integer_fields,
        'float_filter_instance': float_filter_instance,
        'float_fields': unique_model.float_fields,
        'encrypted_fields': unique_model.encrypted_fields if hasattr(unique_model, 'encrypted_fields') else [],
        'unique_id': unique_id,
        'unique_model_name': unique_model_name,
        'unique_field_name': unique_field_name,
        'unique_value': unique_value,
        'unique_filter': list(set(user_unique_filter + date_unique_filter + int_unique_filter + float_unique_filter)),
        'table_name': unique_model_name,
        'fields': fields,
        'saved_filters': saved_filters_json,

        # snapshots
        'snapshots': snapshots,
        'latest_snapshot': latest_snapshot,
        'selected_snapshot': summary
    }
    
    return render(request, 'apps/unique_details.html', context)

# Not Required
# @login_required(login_url='/users/signin/')
# def delete_unique_details(request, id):
#     software = DeviceView.objects.get(pk=id)
#     software.delete()
#     return redirect(request.META.get('HTTP_REFERER'))

# Not Required
# @login_required(login_url='/users/signin/')
# def update_unique_details(request, id):
#     software = DeviceView.objects.get(pk=id)

#     if request.method == 'POST':
#         for attribute, value in request.POST.items():
#             if attribute == 'csrfmiddlewaretoken':
#                 continue

#             if value == '':
#                 if attribute in DeviceView.integer_fields:
#                     value = None
#                 elif attribute in DeviceView.float_fields:
#                     value = None
#                 else:
#                     continue
            
#             setattr(software, 'Total_Contract_Value_Per_Month_Excluding_VAT', calculate_contract_value(request))
#             setattr(software, 'Contract_Remainder_In_Days', calculate_remaining_days(request))
           

#             if not attribute in DeviceView.date_fields_to_convert:
#                 setattr(software, attribute, value)
#             else:
#                 unix_time = setattr(software, attribute, value)
#                 possible_formats = ['%Y-%m-%dT%H:%M:%S', '%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M:%S', '%Y-%m-%d']
#                 for format_str in possible_formats:
#                     try:
#                         if value:
#                             unix_time = datetime.strptime(value, format_str).timestamp()
#                             setattr(software, attribute, unix_time)
#                             break
#                     except ValueError:
#                         pass 

#         software.save()

#     return redirect(request.META.get('HTTP_REFERER'))


# Export as CSV
class ExportCSVView(View):
    def get(self, request, unique_model_name, unique_field_name, unique_value):
        unique_id = f"{unique_model_name}.{unique_field_name}"
        model = apps.get_model('tables', unique_model_name)
        unique_details = model.objects.filter(**{unique_field_name: unique_value})

        db_field_names = [field.name for field in model._meta.get_fields() if not field.is_relation]

        # Snapshot
        latest_snapshot, summary, snapshot_filter, snapshots = common_snapshot_filter(request, unique_model_name)

        fields = []
        pre_column = ('ID', 'loader_instance','hash_data', 'json_data', )

        show_fields = HideShowFilter.objects.filter(value=False, userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id)
        for field in show_fields:
            if field.key not in pre_column:
                fields.append(field.key)

        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = f'attachment; filename="{unique_model_name}_unique_details.csv"'       

        filter_string = {}
        filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.TEXT)
        combined_q_objects, user_unique_filter, user_query_conditions, user_count_filters = same_key_filter(filter_instance, return_count_filters=True)

        # for date range
        date_filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.DATE)
        date_string, date_unique_filter, date_query_conditions, date_count_filters = common_date_filter(date_filter_instance, return_count_filters=True)
        filter_string.update(date_string)
        
        # for integer range
        int_filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.INT)
        int_string, int_unique_filter, int_query_conditions, int_count_filters = common_integer_filter(int_filter_instance, return_count_filters=True)
        filter_string.update(int_string)

        # for float range
        float_filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.FLOAT)
        float_string, float_unique_filter, float_query_conditions, float_count_filters = common_float_filter(float_filter_instance, return_count_filters=True)
        filter_string.update(float_string)

        order_by = request.GET.get('order_by', 'pk')
        base_queryset = unique_details.filter(**filter_string).filter(combined_q_objects).filter(**snapshot_filter)
        queryset = base_queryset

        if user_query_conditions:
            queryset = queryset.filter(user_query_conditions)
        if date_query_conditions:
            queryset = queryset.filter(date_query_conditions)
        if int_query_conditions:
            queryset = queryset.filter(int_query_conditions)
        if float_query_conditions:
            queryset = queryset.filter(float_query_conditions)
        
        if user_count_filters:
            queryset = common_count_filter(user_count_filters, base_queryset, queryset, fields)
        elif date_count_filters:
            queryset = common_count_filter(date_count_filters, base_queryset, queryset, fields)
        elif int_count_filters:
            queryset = common_count_filter(int_count_filters, base_queryset, queryset, fields)
        elif float_count_filters:
            queryset = common_count_filter(float_count_filters, base_queryset, queryset, fields)
        else:
            if order_by == 'count' or order_by == '-count':
                order_by = 'pk'
            if 'count' in fields:
                fields.remove('count')

        queryset = queryset.order_by(order_by)

        if user_unique_filter:
            queryset = common_unique_filter(request, user_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')
        if date_unique_filter:
            queryset = common_unique_filter(request, date_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')
        if int_unique_filter:
            queryset = common_unique_filter(request, int_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')
        if float_unique_filter:
            queryset = common_unique_filter(request, float_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')

        softwares = software_filter(request, queryset, db_field_names)

        writer = csv.writer(response)
        writer.writerow(fields) 

        for software in softwares:
            row_data = []
            for field in fields:
                if field in model.date_fields_to_convert:
                    unix_timestamp = getattr(software, field)
                    if unix_timestamp:
                        date_time = datetime.utcfromtimestamp(unix_timestamp)
                        formatted_date = date_time.strftime('%b %d %Y %I:%M%p')
                        row_data.append(formatted_date)
                    else: 
                        row_data.append("")
                else:
                    row_data.append(getattr(software, field))

            writer.writerow(row_data)

        return response


class ExportExcelView(View):
    def get(self, request, unique_model_name, unique_field_name, unique_value):
        unique_id = f"{unique_model_name}.{unique_field_name}"
        model = apps.get_model('tables', unique_model_name)
        unique_details = model.objects.filter(**{unique_field_name: unique_value})

        db_field_names = [field.name for field in model._meta.get_fields() if not field.is_relation]

        # Snapshot
        latest_snapshot, summary, snapshot_filter, snapshots = common_snapshot_filter(request, unique_model_name)

        fields = []
        pre_column = ('ID', 'loader_instance', 'hash_data', 'json_data', )

        show_fields = HideShowFilter.objects.filter(
            value=False, 
            userID=get_user_id(request), 
            parent=ModelChoices.UNIQUE, 
            unique_id=unique_id
        )
        for field in show_fields:
            if field.key not in pre_column:
                fields.append(field.key)

        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{unique_model_name}_unique_details.xlsx"'

        filter_string = {}
        filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.TEXT)
        combined_q_objects, user_unique_filter, user_query_conditions, user_count_filters = same_key_filter(filter_instance, return_count_filters=True)

        # date filter
        date_filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.DATE)
        date_string, date_unique_filter, date_query_conditions, date_count_filters = common_date_filter(date_filter_instance, return_count_filters=True)
        filter_string.update(date_string)
        
        # int filter
        int_filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.INT)
        int_string, int_unique_filter, int_query_conditions, int_count_filters = common_integer_filter(int_filter_instance, return_count_filters=True)
        filter_string.update(int_string)

        # float filter
        float_filter_instance = SavedFilter.objects.filter(userID=get_user_id(request), parent=ModelChoices.UNIQUE, unique_id=unique_id, field_type=FieldType.FLOAT)
        float_string, float_unique_filter, float_query_conditions, float_count_filters = common_float_filter(float_filter_instance, return_count_filters=True)
        filter_string.update(float_string)

        order_by = request.GET.get('order_by', 'pk')
        base_queryset = unique_details.filter(**filter_string).filter(combined_q_objects).filter(**snapshot_filter)
        queryset = base_queryset

        if user_query_conditions:
            queryset = queryset.filter(user_query_conditions)
        if date_query_conditions:
            queryset = queryset.filter(date_query_conditions)
        if int_query_conditions:
            queryset = queryset.filter(int_query_conditions)
        if float_query_conditions:
            queryset = queryset.filter(float_query_conditions)
        
        if user_count_filters:
            queryset = common_count_filter(user_count_filters, base_queryset, queryset, fields)
        elif date_count_filters:
            queryset = common_count_filter(date_count_filters, base_queryset, queryset, fields)
        elif int_count_filters:
            queryset = common_count_filter(int_count_filters, base_queryset, queryset, fields)
        elif float_count_filters:
            queryset = common_count_filter(float_count_filters, base_queryset, queryset, fields)
        else:
            if order_by == 'count' or order_by == '-count':
                order_by = 'pk'
            if 'count' in fields:
                fields.remove('count')

        queryset = queryset.order_by(order_by)

        if user_unique_filter:
            queryset = common_unique_filter(request, user_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')
        if date_unique_filter:
            queryset = common_unique_filter(request, date_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')
        if int_unique_filter:
            queryset = common_unique_filter(request, int_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')
        if float_unique_filter:
            queryset = common_unique_filter(request, float_unique_filter, queryset, snapshot_filter, f'tables_{unique_model_name}')

        softwares = software_filter(request, queryset, db_field_names)

        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"{unique_model_name}_details"

        # Write headers
        for col_num, field in enumerate(fields, 1):
            col_letter = get_column_letter(col_num)
            ws[f"{col_letter}1"] = field

        # Write data rows
        for row_num, software in enumerate(softwares, 2):
            row_data = []
            for field in fields:
                if field in getattr(model, "date_fields_to_convert", []):
                    unix_timestamp = getattr(software, field)
                    if unix_timestamp:
                        date_time = datetime.utcfromtimestamp(unix_timestamp)
                        formatted_date = date_time.strftime('%b %d %Y %I:%M%p')
                        row_data.append(formatted_date)
                    else: 
                        row_data.append("")
                else:
                    row_data.append(getattr(software, field))
            
            for col_num, value in enumerate(row_data, 1):
                ws.cell(row=row_num, column=col_num, value=value)

        wb.save(response)
        return response

# Not Required
# from reportlab.pdfgen import canvas
# from io import BytesIO
# Export as PDF
# class ExportPDFView(View):
#     def get(self, request):
#         buffer = BytesIO()
#         response_pdf = HttpResponse(content_type='application/pdf')
#         response_pdf['Content-Disposition'] = 'attachment; filename="unique_details.pdf"'

#         pdf = canvas.Canvas(buffer)

#         page_number = request.GET.get('page', 1)
#         items_per_page = 25

#         softwares = DeviceView.objects.all()
#         paginator = Paginator(softwares, items_per_page)
#         current_page_devices = paginator.get_page(page_number)

#         y_position = 800

#         for device in current_page_devices:
#             pdf.drawString(100, y_position, f"ID: {device.ID}, DeviceName: {device.DeviceName}, OperatingSystem: {device.OperatingSystem}")
#             y_position -= 20

#         pdf.save()

#         buffer.seek(0)
#         response_pdf.write(buffer.getvalue())

#         return response_pdf
